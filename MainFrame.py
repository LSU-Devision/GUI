import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tensorflow as tf
import os.path
import numpy as np
from csbdeep.utils import normalize
from PIL import Image
from stardist.models import StarDist2D
from tkinter import messagebox
import csv
import datetime
import time
import threading
import matplotlib
matplotlib.use('agg')
import matplotlib.pyplot as plt
from stardist import random_label_cmap
from Slideshow import Slideshow
import Settings
import Utilities as utils
import CSVEditor as csv_editor
from openpyxl import Workbook, load_workbook
from SettingsWindow import SettingsWindow
from Predictions import Predictions
from ExcelWindow import ExcelWindow
'''
Class Main Frame
Author: Max
Contributors: Skylar Wilson, Alex Mensen-Johnson
Class: Main Frame
Description: Main Frame of GUI, all sub frames will be loaded inside of this class
Params:
    container: a container containing the title and geometry of the Graphic User Interface
Methods:
    init: Initialization method
    create display: method to create the buttons and subframes of the main frame
    load display: method to load data into the GUI
    select files: method for selecting files to be predicted on
    select model: method for selecting the model for prediction
    predict: method for predicting egg count for the frog eggs. outputs the photos saved with the number count
    predict all: predicts the values of all the loaded files in the GUI
    export to csv: Outputs the files to csv
    clear images: Clears the images from the GUI
    help page: creates a help page pop up for the users
'''
class MainFrame(ttk.Frame):
    '''
    method: Initialization method
    Authors: Alex Mensen-Johnson, Skylar Wilson
    '''
    def __init__(self, container):
        super().__init__(container)
        self.settings = Settings.SettingsJson()
        self.container = container
        self.image_files = []
        self.prediction_files = {}
        # initialize predictions_data list -skylar
        self.predictions_data = []
        # initialize lbl_cmap for random color map -skylar
        self.lbl_cmap = random_label_cmap()
        self.model = None
        self.csv_editor = csv_editor.CSVEditor()
        self.csv_editor.set_csv_file(None)
        self.csv_editor.set_csv_label(None)
        self.csv_label_index = 0
        self.is_csv_save_page_open = False
        self.is_settings_page_open = False
        self.device = '/GPU:0' if tf.config.experimental.list_physical_devices('GPU') else '/CPU:0'
        self.progress_popup = None  # Initialize progress_popup to None

        # settings for automatic csv export
        # i made it so it can return a boolean value
        self.automatic_csv_setting = self.settings.get_automatic_csv_export()
        # print(self.automatic_csv_setting)
        self.predict_index = 1


        # settings for automatic prediction data clear
        self.automatic_prediction_data_clear_setting = self.settings.get_automatic_prediction_clear_data()
        # settings for clear data on clear images toggle
        self.clear_data_on_clear_images_setting = self.settings.get_clear_data_on_clear_images()
        # settings for save images output toggle -skylar,
        self.save_images_output_setting = self.settings.get_save_images_output()

        # Create an instance of the Predictions class
        self.predictions = Predictions(self.image_files, self, self.model, self)

        self.create_display()
        self.load_display()

    '''
    Author: Alex Mensen-Johnson
    Creates the display in the code and organizes all the creation into a method
    If you intend on creating a feature for display, add it here please
    '''
    def create_display(self):
        # Creates the select files button
        self.select_files_button = ttk.Button(self, text='Select Files', command=self.select_files)
        # Creates the slide show
        self.slideshow = Slideshow(self)
        # Creates the select model button
        self.select_model_button = ttk.Button(self, text='Select Model', command=self.select_model)
        # Creates the model label
        self.model_label = ttk.Label(self, text='2D_demo')
        # initialize buttons as disabled until a model is selected
        self.model_selected = False
        # Creates the predict all button
        self.predict_all_button = ttk.Button(self, text='Predict All', command=self.predictions.predict_all, state=tk.DISABLED)
        # make two buttons
        self.select_model_button = ttk.Button(self, text='Select Model', command=self.select_model)
        # creates the model label
        self.model_label = ttk.Label(self, text='No Model Selected',font=50)
        # creates clear button
        self.clear_button = ttk.Button(self, text='Clear Images', command=self.clear_images)
        # creates a help button that will display button usage
        self.show_info = ttk.Button(self, text='Help', command=self.help_page)
        # creates the csv label
        self.csv_label_title = ttk.Label(self, text=str(self.csv_editor.get_csv_label()),font=50)
        # creates the export to csv button
        self.excel_window_button = ttk.Button(self, text='Excel Window', command= lambda : self.open_excel_window() )
        # Create settings page
        self.settings_page_button = ttk.Button(self, text='Settings', command=lambda: SettingsWindow(master=self, container=self.container, settings=self.settings))



        # Create the progress bar
        # Create a label to display progress of predicted images
        # Create a label to display estimated time remaining -skylar
        # self.progress_bar = ttk.Progressbar(self, orient='horizontal', mode='determinate', length=300)
        # self.predicted_images_label = ttk.Label(self, text='Predicted 0/0 images')
        # self.estimated_time_label = ttk.Label(self, text='Estimated time remaining: N/A')

    '''
    Author: Alex Mensen-Johnson
    organizes the display buttons into a method for loading the display in a clear format
    If you are loading anything into the display, please add it here.
    '''
    def load_display(self):
        # using grid layout -skylar
        self.grid_rowconfigure(0, weight=0)
        self.grid_columnconfigure(0, weight=1)

        # loads the model label into the frame
        self.model_label.grid(row=1, column=0, pady=5)
        # loads the select model button into the frame
        self.select_model_button.grid(row=2,column=0,pady=5)
        # Loads the select files button into the page
        self.select_files_button.grid(row=3, column=0, pady=5)
        # loads the predict all button
        self.predict_all_button.grid(row=4,column=0,pady=5)
        # loads the slide show frame into the display
        self.slideshow.grid(row=5, column=0)
        # adds clear button using grid method
        self.clear_button.grid(row=6, column=0, pady=5)
        # adds the button to the GUI
        self.show_info.grid(row=7, column=0, pady=5)
        # adds the csv label to the window
        self.csv_label_title.grid(row=8, column=0, pady=0)
        # adds the csv save page to the window
        self.excel_window_button.grid(row=9, column=0, pady=5)
        # adds the settings button to the window
        self.settings_page_button.grid(row=10,column=0,pady=5)
        # Add the progress bar and labels to the window -skylar
        # self.progress_bar.grid(row=11, column=0, pady=5)
        # self.predicted_images_label.grid(row=12, column=0, pady=5)
        # self.estimated_time_label.grid(row=13, column=0, pady=5)

    '''****************************************************************************************'''
    '''Destroys current frame and reloads MainFrame for the purpose of dynamic display. -skylar'''
    def update_display(self):
        parent_container = self.container
        self.destroy()
        
        # Re-initialize the frame with the same container
        new_frame = MainFrame(parent_container)

        # Add the new frame to the container using grid
        new_frame.grid(row=0, column=0, sticky='nsew')

    def disable_button(self, button):
        button.config(state = tk.DISABLED)

    def enable_button(self, button):
        button.config(state = tk.NORMAL)

    def select_files(self):
        files = filedialog.askopenfilenames(initialdir='/home/max/development/stardist/data')
        self.image_files.extend(files)
        self.predictions.image_files = self.image_files
        if self.image_files:
            self.slideshow.update_image()

    def select_model(self):
        model_path = filedialog.askdirectory()
        print(model_path)
        print(os.path.basename(model_path))
        print(os.path.dirname(model_path))
        with tf.device(self.device):
            self.model = StarDist2D(None, name=os.path.basename(model_path), basedir=os.path.dirname(model_path))
        self.predictions.model = self.model
        self.model_label.config(text=os.path.basename(model_path))
        if not self.model_selected:
            self.predict_all_button.config(state=tk.NORMAL)
            self.model_selected = True

    '''
    function to export filenames, predicted counts, and date/time to a csv file in the
    output folder -skylar
    '''
    def export_predictions_to_csv(self):
        current_time = datetime.datetime.now().strftime("%m-%d-%Y_%I-%M-%S %p")
        if self.csv_editor.get_csv_file() is None:
            self.csv_editor.set_csv_file(os.path.join('output', f'predictions_{current_time}.xlsx'))
        if not os.path.exists('output'):
            os.makedirs('output')

        #with open(self.csv_editor.get_csv_file(), mode='w', newline='') as file:
            #writer = csv.writer(file)
            #writer.writerow(['Index', 'Date', 'Time', 'File Name', ' Total Count'])
            #writer.writerows(self.predictions_data)
        does_file_exist = True
        try:
            wb = load_workbook(self.csv_editor.get_csv_file())
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            does_file_exist = False
        has_index_column = 'Index' in [cell.value for cell in ws['A']]
        last_index_value = 0
        if does_file_exist is False:
            headers = ['Index', 'Date', 'Time', 'File Name', ' Total Count']
            ws.append(headers)
        elif has_index_column is True:
            for cell in ws['A']:
                last_index_value = cell.value
        for prediction in self.predictions.predictions_data:
            prediction_list = list(prediction)
            prediction_list[0] = prediction_list[0] + last_index_value
            ws.append(prediction_list)
        try:
            wb.save(self.csv_editor.get_csv_file())
        except PermissionError:
            messagebox.showerror("Permissions Error", "Excel file may have Read only attribute, please check permissions")
        if self.settings.get_automatic_prediction_clear_data():
            self.predictions_data.clear()

    '''
    Clear Images: By Alex Mensen-Johnson
    Method for clearing Images from the GUI, resets all variables to their primary state.
    
    '''
    def clear_images(self):
        # Set the base image reference to None
        self.slideshow.base_image.image_ref = None
        # Clear the label of the file path above the slideshow
        self.slideshow.filepath_label.config(text='')
        # Set the index of the pictures to 0
        self.slideshow.current_index = 0
        # empty the image files array
        self.image_files = []
        # set the image to none
        self.slideshow.base_image.set_image(None)
        # set the image files to the current image files, which is none
        self.slideshow.image_files = self.image_files
        # set the predicted image to none
        self.slideshow.predicted_image.set_image(None)
        # set the item count label to empty
        self.slideshow.item_count_label.config(text=' ')
        # resets the predictions file to the empty dictionary
        self.predictions.prediction_files = {}

        # sets the prediction files to empty
        self.prediction_files = {}
        # sets the slideshow prediction files to empty
        self.slideshow.prediction_files = self.prediction_files

        self.slideshow.current_index = 0

        if self.settings.get_clear_data_on_clear_images():
            self.predictions.predictions_data.clear()
            self.predictions_data.clear()
        #
        messagebox.showinfo("Devision", "Images Cleared!")

    def load_csv_by_selection(self):
        csv_file = filedialog.askopenfilename(initialdir='/home/max/development/stardist/data')
        self.csv_editor.set_csv_file(csv_file)
        self.csv_editor.get_substring()
        self.csv_label_title.config(text= str(self.csv_editor.get_csv_label()))

    '''
    Help Page: by Alex Mensen-Johnson
    Description: Loads the Help_Information.txt into a file, reads the file, and displays the information in a pop up
    '''
    def help_page(self):
        # Load File
        info_file = open("docs/Help_Information.txt")
        # Read the file
        file_information = info_file.read()
        # Create the title string
        title = 'Help'
        # Create pop up with the information
        messagebox.showinfo(title,file_information)

    def clear_predicted_data(self):
        self.predictions.predictions_data.clear()

    def clear_csv_file(self):
        self.csv_editor.set_csv_file(None)
        self.csv_editor.set_csv_label(None)
        self.csv_label_title.config(text='None')
    # change test added here

    '''
    function to open the excel window
    '''
    def open_excel_window(self):
        # check if the excel window is already open
        if self.is_csv_save_page_open == False:
            # open the excel window
            ExcelWindow(master=self, container=self.container, excel_editor=self.csv_editor)

