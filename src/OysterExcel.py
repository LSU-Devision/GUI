import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from scipy.stats import t
from pathlib import Path
import os

if not os.path.exists('excel'):
    os.mkdir('excel')

class OysterExcel():
    id = 0
    def __init__(self, *, file_name='oyster-data.xlsx', staff_name=''):
        self.id = OysterExcel.id 
        OysterExcel.id += 1
        
        self.file_name = file_name
        
        formatted_datetime = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
        
        # Goes into the info tab
        self.info_df = pd.DataFrame([[formatted_datetime, staff_name]], columns=['Date', 'Staff'])
        
        # Goes into the data tab
        # Mass is in grams unless otherwise specified
        self.df = pd.DataFrame(columns=[
            'model',
            'group',
            'file-name',
            'size-class',
            'seed-tray-weight', 
            'slide-weight', 
            'slide-and-seed-weight', 
            'subsample-count', 
            'total-number'
        ])
        
        self.data_to_readable = {
            'model':'Model',
            'group':'Group Number',
            'file-name': 'File Name',
            'size-class': 'Size Class',
            'seed-tray-weight': 'Seed Tray Weight (g)',
            'slide-weight': 'Slide Weight (g)',
            'slide-and-seed-weight': 'Slide and Seed Weight (g)',
            'subsample-count': 'Subsample Count',
            'total-number': 'Total Number'}
        
        self.readable_to_data = {self.data_to_readable[k]:k for k in self.data_to_readable}
        
        # Goes into the statistics tab
        self.stats = None
        
    def insert(self, *, model, group_number, file_name, size_class,
                        seed_tray_weight, slide_weight,
                        slide_and_seed_weight, subsample_count):
        """Inserts a new value into the OysterExcel dataframe

        Args:
            model (str): The name of the model used to predict the sample
            group_number (int): The group number of the group being inserted
            file_name (str): The file name where the group count was taken from
            size_class (str): The size class of the group being inserted
            seed_tray_weight (float): The weight of the seed tray used on the group being inserted in grams
            slide_weight (float): The weight of the slide of the group being inserted in grams
            slide_and_seed_weight (float): The weight of the slide and the weight of the seed of the group being inserted in grams
            subsample_count (int): The predicted number of oysters in this sample
        """
        
        
        total_count = (subsample_count / (slide_and_seed_weight - slide_weight)) * seed_tray_weight
        insert_row = pd.DataFrame([[
                model,
                group_number,
                file_name, 
                size_class,
                seed_tray_weight, 
                slide_weight, 
                slide_and_seed_weight,
                subsample_count,
                total_count
            ]], columns=self.df.columns
        )
        
        
        self.df = pd.concat([self.df, insert_row], ignore_index=True)
        self.compute()
        
    # Takes in a dataframe (of insertable values) and concats it to the current dataframe
    def extend(self, insert_df):
        """Insert all the values in a given dataframe into this objects dataframe

        Args:
            insert_df (pandas.DataFrame): The dataframe being inserted, it must have columns of the same name as this objects dataframe
        """
        
        pd.options.mode.chained_assignment = None

        
        insert_df = insert_df[['model', 'group', 'file-name', 'size-class', 'seed-tray-weight', 'slide-weight', 'slide-and-seed-weight', 'subsample-count']]
        total_count = insert_df['subsample-count'] / (insert_df['slide-and-seed-weight'] - insert_df['slide-weight']) * insert_df['seed-tray-weight']
        insert_df['total-number'] = total_count
        
        if len(self.df) == 0: 
            self.df = insert_df
        else:
            self.df = pd.concat([self.df, insert_df], ignore_index=True)
       
        self.compute()
        pd.options.mode.chained_assignment = 'warn'
        
    def compute(self):
        """Computes mean, standard deviation, standard error, and the deviation of the 95% confidence interval 
           on all groups and saves them into this objects stats attribute as a dataframe
        """
        groups = self.df[['group', 'total-number']].groupby('group')
        
        # Student's T parameters
        dof = groups.size() - 1 # degrees of freedom
        alpha = .05 # also known as p-value
        ppf = t.ppf(1-(alpha/2), dof) # Critical t-value for given degrees of freedom at alpha threshold
        
        # Aggregated statistics
        mean = groups.mean().rename(columns={'total-number':'mean'})
        std = groups.std().rename(columns={'total-number':'std'})
        sem = groups.sem().rename(columns={'total-number':'sem'})
        dof = dof.rename('dof')
        
        confidence95 = sem.apply(lambda x: x*ppf).rename(columns={'sem':'confidence95'})
        
        # Write to dataframe
        self.stats = pd.concat([mean, std, sem, dof, confidence95], axis=1)
    
    def _openpyxl_write(self, dataframe, workbook, worksheet, kwargs):
        """Internal method for formatting openpyxl

        Args:
            dataframe (pandas.DataFrame): The dataframe object that is being written
            workbook (openpyxl.workbook.workbook.Workbook): The excel workbook object
            worksheet (openpyxl.worksheet.worksheet.Worksheet): The current working sheet
            kwargs (dict): Additional keyword arguments to be passed into openpyxl's dataframe_to_rows
        """
        
        workbook.active = worksheet
        # Add rows to the current sheet from the dataframe
        for row in dataframe_to_rows(dataframe, **kwargs):
            worksheet.append(row)
        
        # Resize columns based on the max width of the entries in that column for this sheet
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            max_width = len(str(
                max(column, 
                    key = lambda x: len(str(x.value))
                    )
                .value))

            worksheet.column_dimensions[column_letter].width = max_width + 1
    
    def write_excel(self):
        """Writes this object into an excel file with three sheets:
           info, which contains date and staff fields;
           data, which contains data of all subsamples;
           and statistics, which contains aggregate data performed on subsamples
        """
        file_path = Path('excel') / Path(f'data{self.id}.xlsx')
        
        # Converting the internal dataframe names to human readable export
        print_df = self.df.copy()
        print_df = print_df.rename(columns=self.data_to_readable)
        print_df.index.name = 'Index'
        
        print_stats = self.stats.copy()
        print_stats = print_stats.rename(columns={
            'mean':'Mean',
            'std':'Standard Deviation',
            'sem':'Standard Error',
            'dof':'Degrees of Freedom (n-1)',
            'confidence95':'95% Confidence Interval'
        })
        print_stats.index.name = 'Group Number'
        
        # Writing to excel using openpyxl
        wb = Workbook()
        
        # Sets the default sheet to the info sheet
        info = wb.active
        data = wb.create_sheet(title="Data")
        statistics = wb.create_sheet(title="Statistics")
        
        info.title = "Info"
        
        self._openpyxl_write(self.info_df, wb, info, {'index':False, 'header':True})

        self._openpyxl_write(print_df, wb, data, {'index':True, 'header':True})

        self._openpyxl_write(print_stats, wb, statistics, {'index':True, 'header':True})
        
        wb.save(file_path)
    
    #This assumes that the file was written by this excel writer or is in the same format
    def read_excel(self, file_path=None):
        if not file_path:
            file_path = self.file_name
            
        df = pd.read_excel(file_path, index_col=0, skiprows=[1], sheet_name='Data')
        df = df.rename(columns=self.readable_to_data)
        
        self.extend(df)
        
        return df

# Testing function
if __name__ == '__main__':  
    import_df = pd.read_csv('test/oyster-example-data.csv')
    excel_obj = OysterExcel(file_name='test/oyster-excel-out.xlsx')
    
    excel_obj.extend(import_df)
    excel_obj.write_excel()
    
    print(excel_obj.read_excel())
    
