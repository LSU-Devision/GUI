from tkinter import messagebox, filedialog
import openpyxl
import Utilities as utils
class OysterPage:
    '''
    class: OysterPage
    description: class for the oyster page. currently handles exports and calculations for the oyster page.
    future implementations: modify to create entire oyster Excel file, needs testing and handling for special cases,
        special cases include, mismatching column names, not one to one set of data, not onto data sets, overwriting versus blank data
        also needs to include user warnings for empty data, incorrect file selection, data overwriting.
        far from complete
    modularity: later on, optimize methods for efficiency, certain algorithms have time complexity of n^4. fix this.
    '''
    def __init__(self,master):
        self.master = master
        self.prediction_files = master.predictions.prediction_files
        self.excel_file = None
        #stores the row index of the column headers
        self.index_row = None
        #stores the cells of the index row, i.e the different assigned column values
        self.index_row_cells = None
        #stores the column headers as the key and the cells as the value
        self.index_row_dict = {}
        # stores the row values of the file names as the key and the predicted count as the value
        self.results_dict = {}

    def calculate(self,sample_weight,subsample_weight,predicted_number):
        '''
        method: calculate
        description: calculates the predicted total count from the given user input
        '''
        subsamples_per_sample = sample_weight / subsample_weight
        predicted_total_count = subsamples_per_sample * predicted_number
        return predicted_total_count

    def load_excel_file(self):
        '''
        method: load_excel_file
        description: loads the excel file from user input
        '''
        self.excel_file = filedialog.askopenfilename(initialdir='/home/max/development/stardist/data')
        if utils.is_excel_file(self.excel_file) is False:
            self.excel_file = None
            messagebox.showerror('File Error', 'File is not an excel file, available file types are .xlsx, .xls, .xlsm, .xltm, .xltx, .xls,.csv')
        self.master.oyster_excel_label.config(text= str(utils.string_to_substring(self.excel_file)))

    def check_for_excel_file(self):
        '''
        method: check_for_excel_file
        description: checks if an excel file has been loaded
        '''
        if self.excel_file is None:
            return False
        return True
    def oyster_excel_reader(self):
        '''
        method: oyster_excel_reader
        description: reads in the excel file, and extracts key values for the program export,\n
        stores the index row value, the index row cells, and the index row dictionary
        Note: the index row dictionary is a dictionary with the column headers as the key and the cells as the value
        Note: load Excel, and check for Excel file before using this method. both methods are run before this in the run method.
        '''
        # open the the excel file using openpyxl. This does not load the worksheet, only the excel file
        workbook = openpyxl.load_workbook(self.excel_file)
        # set the workbook to the 1st page worksheet
        worksheet = workbook.active
        # gets the rows from the worksheet
        rows = worksheet.rows
        self.index_row = None
        self.index_row_cells = None
        self.index_row_dict = {}
        # for loop has time complexity of n^3, fix this. try index based accessing to make it smaller
        # finds the Treatment or Group name and stores all the cells in that row creating a dictionary
        for row in rows:
            if row[0].value is not None:
                if "Treatment or Group Name" in row[0].value:
                    self.index_row = row[0].row
                    self.index_row_cells = row
                    for cell in self.index_row_cells:
                        column_values = []
                        for col_val in worksheet.iter_rows(min_row=cell.row,min_col=cell.column,max_col=cell.column):
                            column_values.append(col_val)
                        self.index_row_dict[cell.value] = column_values
        workbook.close()
        if len(self.index_row_dict) == 0:
            return False
        return True

    def get_predicted_values(self,test = False):
        '''
        method: get_predicted_values
        description: prints out the predicted values, useful for examining possible errors in the code, also returns true or false if the predictions data is not empty
        note: if you want to enable the predictions data to be printed, set test to true, by default it is false
        '''
        if self.master.predictions.predictions_data is not None and test is True:
            for prediction in self.master.predictions.predictions_data:
                print(f' Prediction file : {prediction[3]} Predicted Number {prediction[4]}')
        if len(self.master.predictions.predictions_data) == 0:
            return False
        return True

    def match_predicted_values(self):
        '''
        method: match_predicted_values
        description: matches the predicted values given the existing filenames in the Excel file\n
        needs to be tested to see how it handles mismatches between the predicted values and the file names
        note: file names can be loaded with or without quotations, for example "filename" will match filename
        '''
        # empties the results dictionary
        self.results_dict = {}
        # Loops through the filenames column to create a dictionary matching row index's to predicted values
        # time complexity is n^3 because of the nested for loops fix this possibly by index based accessing
        excel_file_name_count = -1 # Initialized at -1 to account for the header
        prediction_count = len(self.master.predictions.predictions_data)
        matched_file_name_count = 0
        for array in self.index_row_dict['File Names']:
            # loops through the array
            for value in array:
                if value.value is None:
                    continue
                else:
                    excel_file_name_count += 1
                for prediction in self.master.predictions.predictions_data:
                    temp_file_name = value.value.split('\\')[-1].replace('"','')
                    if temp_file_name == prediction[3]:
                        self.results_dict[value.row] = prediction[4]
                        matched_file_name_count += 1
        if excel_file_name_count != prediction_count or matched_file_name_count != excel_file_name_count:
            return False
        return True

    def check_for_column(self,column_name):
        if column_name in self.index_row_dict:
            return True
        return False
    def export_data_into_sheet(self):
        # open the workbook
        workbook = openpyxl.load_workbook(self.excel_file)
        # set the workbook to the 1st page worksheet
        worksheet = workbook.active
        # creates a target column store for accessing the cells by column, result should contain the column of the Subsample count
        target_column = None
        # access the Subsample count column
        # despite nested array, really is n = 2, as noted by the break statements. can be optimized for more efficent access, do later
        for cell_array in self.index_row_dict['Subsample Count']:
            # for loop to access cells to retrieve column letter
            for cell in cell_array:
                target_column = cell.column_letter
                # breaks as for loop is only used once
                break
            # breaks as for loop is only used once
            break
        # uncomment for column checking
        # print(target_column)
        # for loop to write results to dictionary by row
        for key in self.results_dict:
            worksheet[f'{target_column}{key}'] = self.results_dict[key]
        # save the workbook
        workbook.save(self.excel_file)
        # close the workbook.
        workbook.close()


    def run_export_methods(self):
        '''
        method: run_export_methods
        description: modular method, runs the methods to export the data
        note: each method is run in an if statement, which returns either true or false. This triggers a warning message or proceeds to the next method.
        '''
        if self.check_for_excel_file() is False:
            messagebox.showerror('File Error','Export Failed, no excel file loaded into program')
            return

        if self.oyster_excel_reader() is False:
            messagebox.showerror("Column naming Error", "Export Failed, 'No Treatment or Group Name' column not found in the excel file. Please be mindful of spelling, capitalization, and spacing errors.")
            return

        if self.get_predicted_values() is False:
            messagebox.showerror('Prediction Error','Export Failed, no predictions found, Export will not be successful')
            return
        if self.check_for_column('File Names') is False:
            messagebox.showerror('File Names Error','Export Failed, no File Names column found in the excel file. Please be mindful of spelling, capitalization, and spacing errors.')
            return
        if self.match_predicted_values() is False:
            messagebox.showerror('Prediction Error','Export Failed, mismatch between predicted values and file names, currently not handling partial exports')
            return

        if self.check_for_column('Subsample Count') is False:
            messagebox.showerror('Subsample Count Error','Export Failed, no Subsample Count column found in the excel file. Please be mindful of spelling, capitalization, and spacing errors.')
            return
        try:
            self.export_data_into_sheet()
        except PermissionError:
            messagebox.showerror('File Error', 'Export Failed, file is open in another program, or File is saved as read only')
            return
        messagebox.showinfo('Export Success', 'Oyster Excel Export was successful')


