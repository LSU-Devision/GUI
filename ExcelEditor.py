import datetime
import json
import os
import Utilities as utils
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

# load default settings
default_settings = 'config/excel_settings.json'
'''
Class: ExcelEditor
Contributors: Alex Mensen-Johnson, Skylar Wilson
Description: Class for the Excel Editor to manage excel files for the master
Params:
    master = master class aka MainFrame
methods:
    __init__: initialize the class
    load_excel_settings: load the excel settings
    save_excel_settings: save the excel settings
    export_predictions_to_excel: export the predictions to excel
    various getters and setter methods
'''
class ExcelEditor:
    '''
    method: __init__
    description: initialize the class
    '''
    def __init__(self,master=None):
        # set the Excel file to none
        self.excel_file = None
        # set the Excel label to none
        self.excel_label = None
        # set the Excel index value to none
        self.excel_index_value = 'None'
        # set the Excel date value to none
        self.excel_date_value = 'None'
        # set the Excel time value to none
        self.excel_time_value = 'None'
        # set the Excel file name value to none
        self.excel_file_name_value = 'None'
        # set the Excel total count value to none
        self.excel_total_count_value = 'None'
        # create output folder variable
        self.output_folder = 'output'
        # set the master to the master
        self.master = master
        # load the Excel settings
        self.load_excel_settings()

    '''
    method: load_excel_settings
    description: load the excel settings in to the class
    '''
    def load_excel_settings(self):
        # create the self.data variable
        self.data = None
        # open the default settings
        with open(utils.resource_path(default_settings)) as json_file:
            # load the default settings from the json file
            self.data = json.load(json_file)
        # set the Excel index value
        self.excel_index_value = self.data['excel_index_column']
        # set the Excel date value
        self.excel_date_value = self.data['excel_date_column']
        # set the Excel time value
        self.excel_time_value = self.data['excel_time_column']
        # set the Excel file name value
        self.excel_file_name_value = self.data['excel_file_name_column']
        # set the Excel total count value
        self.excel_total_count_value = self.data['excel_total_count_column']

    '''
    method: save_excel_settings
    description: save the excel settings in to the class
    '''
    def save_excel_settings(self):
        # create the self.data_save variable
        self.data_save = {}
        # set the data save Excel index column
        self.data_save['excel_index_column'] = self.excel_index_value
        # set the data save Excel date column
        self.data_save['excel_date_column'] = self.excel_date_value
        # set the data save Excel time column
        self.data_save['excel_time_column'] = self.excel_time_value
        # set the data save Excel file name column
        self.data_save['excel_file_name_column'] = self.excel_file_name_value
        # set the data save Excel total count column
        self.data_save['excel_total_count_column'] = self.excel_total_count_value
        # open the default settings
        with open(utils.resource_path(default_settings), 'w') as outfile:
            # dump the data to the json file
            json.dump(self.data_save, outfile,indent=4)

    '''
    method: export_predictions_to_excel
    description: export the predictions to excel
    '''
    def export_predictions_to_excel(self):
        # check to see if the data has been cleared
        if self.master.is_data_cleared == False:
            # check to see if the user wants to continue
            no_cancel = messagebox.askokcancel(title='Data Not Cleared',message='Data has not been cleared since last export. Do you want to continue?',parent=self.master)
            # if the user does not want to continue
            if no_cancel == False:
                # return the function
                return
        # set the current time with a format
        current_time = datetime.datetime.now().strftime("%m-%d-%Y_%I-%M-%S %p")
        # check to see if the Excel file exists
        if self.get_excel_file() is None:
            # check to see if a name has been set to the excel_file_variable
            if self.master.excel_editor.get_excel_file() is None:
                # if none has been set, set the Excel file to the default
                self.set_excel_file(os.path.join(self.output_folder, f'predictions_{current_time}.xlsx'))
            else:
                # if a name has been set, set the Excel file to the name
                self.set_excel_file(os.path.join(self.output_folder, f'{self.master.excel_window.get_excel_file_variable()}.xlsx'))
        # check to see if the output folder exists
        if not os.path.exists(self.output_folder):
            # create the output folder
            os.makedirs(self.output_folder)
        # create boolean variable to check if the file exists
        does_file_exist = True
        # create a list to store the headers
        excel_headers = []
        # try to create the workbook
        try:
            # load the workbook
            wb = load_workbook(self.get_excel_file())
            # set the worksheet
            ws = wb.active
            # iterate over the header row in the worksheet
            for cell in ws[1]:
                # append the headers to the list
                excel_headers.append(cell.value)
        # if the file does not exist
        except FileNotFoundError:
            # create the workbook
            wb = Workbook()
            # set the worksheet
            ws = wb.active
            # set the boolean variable to false
            does_file_exist = False
        # create a boolean variable to check if the index column exists
        has_index_column = 'Index' in excel_headers
        # set a value for the last index value
        last_index_value = 0
        # if the file does not exist
        if does_file_exist is False:
            # set the headers to the first row
            ws.cell(row=1, column=1, value=self.get_excel_headers()[0])
            # append the headers to the worksheet
            for i, header in enumerate(self.get_excel_headers()[1:], start=2):
                # add each header to the appropriate row
                ws.cell(row=1, column=i, value=header)
                # get the index of the headers
            index_list = self.get_headers_index()
            # create an empty list for indexing
            edited_prediction_list = []
            # iterate over the predictions
            for prediction in self.master.predictions.predictions_data:
                # clear the prediction list
                edited_prediction_list.clear()
                # iterate over the index values
                for index in index_list:
                    # append the index value from predictions to the list
                    edited_prediction_list.append(prediction[index])
                # append the list to the worksheet
                ws.append(edited_prediction_list)
                # update the internal excel label with the substring
                self.get_substring()
                # update the Mainframe excel label
                self.master.excel_label_title.config(text=self.get_excel_label())
        # if the index column does exist
        elif has_index_column is True:
            # get the index column position
            index_column_position = excel_headers.index('Index')
            # set the index column letter variable to none, later to be updated with the column letter of the index
            index_column_letter = None
            # iterate over the header row in the worksheet
            for cell in ws[1]:
                # check if the cell value is index
                if cell.value == 'Index':
                    # update the index column letter
                    index_column_letter = cell.column_letter
                    # break the loop
                    break
            # iterate over the index column
            for cell in ws[index_column_letter]:
                # update the last index value to the current cell value, traversing down the Excel until out of values
                last_index_value = cell.value
        # iterate over the predictions to append to the worksheet
        if does_file_exist is True:
            # get the possible headers
            possible_headers = ['Index', 'Date', 'Time', 'File Name', 'Total Count']
            # get the headers from the Excel file
            excel_header_list = []
            # iterate over the headers
            for cell in ws[1]:
                # append the header to the list if it is in the possible headers
                if cell.value in possible_headers:
                    # append the header to the list
                    excel_header_list.append(cell.value)
                # if the header is not in the possible headers
                elif cell.value is None:
                    # show an error
                    messagebox.showerror("Error", f"Invalid headers in excel file. There is an empty cell before the last value in the header row.")
                    # return
                    return
                else:
                    # show an error
                    messagebox.showerror("Error", f"Invalid headers in excel file. This is the invalid header {cell.value}")
                    # return
                    return
            # check to see if the header list is unique
            if len(excel_header_list) != len(set(excel_header_list)):
                # show an error
                messagebox.showerror("Error", "Headers in excel file are not unique. This is not supported.")
                # return
                return
            # check to see if the headers are in the same order as the program headers
            order_check = True
            # get the current header list
            current_header_list = self.get_excel_headers()
            # create an index for header access
            index = 0
            # iterate over the headers
            for header in excel_header_list:
                # check to see if the header is in the correct order
                if header != current_header_list[index]:
                    # set the boolean to false
                    order_check = False
                    # break out of the loop
                    break
                # increment the index
                index += 1
            # bool to check if the headers need to be reordered
            re_order_settings = False
            # if the headers are not in the correct order
            if order_check is False:
                # ask the user if they would like to reorder the settings
                re_order_settings = messagebox.askyesno("Notice", "Headers in excel file are not in the correct order. Would you like to reorder the settings to match the order in the excel file?")
                # create an index of the Excel headers
                excel_header_index = self.get_headers_index(excel_header_list)
                # create an empty list for prediction data
                edited_excel_prediction_list = []
                # iterate over the predictions
                for prediction in self.master.predictions.predictions_data:
                    # clear the prediction list
                    edited_excel_prediction_list.clear()
                    # iterate over the index values
                    for index in excel_header_index:
                        # append the index value from predictions to the list
                        edited_excel_prediction_list.append(prediction[index])
                    # if the index column exists
                    if has_index_column is True:
                        # add the index value if the column exists
                        edited_excel_prediction_list[index_column_position] = edited_excel_prediction_list[index_column_position] + last_index_value
                    # append the list to the worksheet
                    ws.append(edited_excel_prediction_list)
            # if the headers are in the correct order
            else:
                # iterate over the predictions
                excel_header_index = self.get_headers_index()
                # create an empty list for prediction data
                edited_excel_prediction_list = []
                # iterate over the predictions
                for prediction in self.master.predictions.predictions_data:
                    # clear the list for new predictions
                    edited_excel_prediction_list.clear()
                    # update the index values to reflect the values on the sheet
                    for index in excel_header_index:
                        # append the index value from predictions to the list
                        edited_excel_prediction_list.append(prediction[index])
                    # update the index value
                    if has_index_column is True:
                        # add the index value if the column exists
                        edited_excel_prediction_list[index_column_position] = edited_excel_prediction_list[index_column_position] + last_index_value
                    # append the list to the worksheet
                    ws.append(edited_excel_prediction_list)
            # if the user wants to reorder the settings
            if re_order_settings is True:
                # create a dictionary of the headers
                header_dict = {}
                # iterate over the possible headers
                for header in possible_headers:
                    # try to get the index of the header
                    try:
                        # get the value of the index and append to the dictionary, adds plus 1 to shift 0-4 to 1-5
                        header_dict[header] = excel_header_list.index(header) + 1
                    # if the header is not in the Excel file
                    except ValueError:
                        # set the value to None
                        header_dict[header] = 'None'
                # set the index value in the class
                self.excel_index_value = str(header_dict['Index'])
                # set the date value in the class
                self.excel_date_value = str(header_dict['Date'])
                # set the time value in the class
                self.excel_time_value = str(header_dict['Time'])
                # set the file name value in the class
                self.excel_file_name_value = str(header_dict['File Name'])
                # set the total count value in the class
                self.excel_total_count_value = str(header_dict['Total Count'])
                # save the Excel settings
                self.save_excel_settings()
                # update the index dropdown value
                self.master.excel_window.excel_index_column_dropdown.set(self.excel_index_value)
                # update the date dropdown value
                self.master.excel_window.excel_date_column_dropdown.set(self.excel_date_value)
                # update the time dropdown value
                self.master.excel_window.excel_time_column_dropdown.set(self.excel_time_value)
                # update the file name dropdown value
                self.master.excel_window.excel_file_name_column_dropdown.set(self.excel_file_name_value)
                # update the total count dropdown value
                self.master.excel_window.excel_total_count_column_dropdown.set(self.excel_total_count_value)
        # try statement to save the workbook
        try:
            # save the workbook
            wb.save(self.get_excel_file())
            # show success
            messagebox.showinfo("Success", "Successfully exported predictions to excel")
        # if the file is read only
        except PermissionError:
            # show an error
            messagebox.showerror("Permissions Error", "Excel file may be open, or you may have Read only attributes, please close the excel file or check permissions")
            # exit the method
            return
        # boolean check for clearing variables
        self.master.is_data_cleared = False
        # boolean check for automatic prediction clearing
        if self.master.settings.get_automatic_prediction_clear_data():
            # clear the predictions
            self.master.predictions.predictions_data.clear()
            # set the boolean check to True
            self.master.is_data_cleared = True


    def get_excel_index_column_index(self):
        return self.excel_index_value
    def get_excel_date_column_index(self):
        return self.excel_date_value
    def get_excel_time_column_index(self):
        return self.excel_time_value
    def get_excel_file_name_column_index(self):
        return self.excel_file_name_value
    def get_excel_total_count_column_index(self):
        return self.excel_total_count_value
    def get_substring(self):
        self.excel_label = utils.string_to_substring(self.excel_file)
    def get_excel_file(self):
        return self.excel_file
    def get_excel_label(self):
        return self.excel_label
    def get_output_folder(self):
        return self.output_folder

    def set_excel_file(self, excel_file):
        if self.output_folder is None:
            self.set_output_folder('output')

        if excel_file is None:
            self.excel_file = None
        else:
            self.excel_file = os.path.join(self.output_folder, os.path.basename(excel_file).split('.')[0] + '.xlsx')
            self.get_substring()

            print(f'excel file is {self.excel_file} and the label is {self.excel_label}')
            self.master.excel_label_title.config(text=self.excel_label)

    def set_excel_label(self, excel_label):
        self.excel_label = excel_label
    def set_excel_index_column_value(self, excel_index):
        self.excel_index_value = excel_index
    def set_excel_date_column_value(self, excel_date):
        self.excel_date_value = excel_date
    def set_excel_time_column_value(self, excel_time):
        self.excel_time_value = excel_time
    def set_excel_file_name_column_value(self, excel_file_name):
        self.excel_file_name_value = excel_file_name
    def set_excel_total_count_column_value(self, excel_total_count):
        self.excel_total_count_value = excel_total_count

    def set_output_folder(self, output_folder):
        self.output_folder = output_folder

    '''
    method: get_excel_headers
    description: get and order the excel headers according to the settings
    '''
    def get_excel_headers(self):
        # create a list of the header values
        value_list = [
            self.excel_index_value,
            self.excel_date_value,
            self.excel_time_value,
            self.excel_file_name_value,
            self.excel_total_count_value
        ]
        # create a new list to store the values
        new_list = []
        # iterate over the values in the header list
        for value in value_list:
            # if the value is not None, append it to the new list
            if value != 'None':
                # append the value
                new_list.append(str(value))
        # sort the new list by numerical value
        new_list.sort()
        # create a new list to store the index values of the header list
        index_list = []
        # iterate over the new list and store the index values in the index list
        for index in new_list:
            # append the index value
            index_list.append(value_list.index(index))
        # create a list for the headers
        headers_list = []
        # iterate over the index list
        for index in index_list:
            # if the index is for the Index
            if index == 0:
                # append the index to the headers list
                headers_list.append('Index')
                # if the index is for the Date
            elif index == 1:
                # append the date to the headers list
                headers_list.append('Date')
            # if the index is for the Time
            elif index == 2:
                # append the time to the headers list
                headers_list.append('Time')
            # if the index is for the File Name
            elif index == 3:
                # append the file name to the headers list
                headers_list.append('File Name')
            # if the index is for the Total Count
            elif index == 4:
                # append the total count to the headers list
                headers_list.append('Total Count')
        # return the headers list
        return headers_list
    '''
    method: get_headers_index
    description: get the index values of the headers
    '''
    def get_headers_index(self,header_list=None):
        # create a list for the index values
        index_list = []
        # get the header list
        if header_list is None:
            header_list = self.get_excel_headers()
        # iterate over the header list
        for header in header_list:
            # if the header is Index
            if header == 'Index':
                # append the index to the index list
                index_list.append(0)
            # if the header is Date
            elif header == 'Date':
                # append the date to the index list
                index_list.append(1)
            # if the header is Time
            elif header == 'Time':
                # append the time to the index list
                index_list.append(2)
            # if the header is File Name
            elif header == 'File Name':
                # append the file name to the index list
                index_list.append(3)
            # if the header is Total Count
            elif header == 'Total Count':
                # append the total count to the index list
                index_list.append(4)
        # return the index list
        return index_list